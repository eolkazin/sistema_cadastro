import customtkinter as ctk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os
from datetime import datetime

# Lista de alimentos e acréscimos com preços
alimentos = [
    "Hambúrguer - R$6,50",
    "Joelho de moça - R$6,50",
    "Coxinha - R$6,50",
    "Kibe - R$6,50",
    "Empada - R$6,50",
    "Risole de milho - R$6,50",
    "Cigarrete de frango com presunto - R$6,50",
    "Tortinha - R$6,50",
    "Pastel assado - R$6,50",
    "Pão de queijo - R$2,50",
    "Café - R$1,50",
    "Café + leite - R$2,50",
    "Toddy P - R$3,00",
    "Toddy G - R$4,00",
    "Vitamina 300ml - R$5,00",
    "Vitamina 400ml - R$6,00",
    "Vitamina 500ml - R$7,00",
    "Vitamina 700ml - R$10,00",
    "Creme de açaí 300ml (3 acompanhamentos) - R$13,00",
    "Creme de açaí 400ml (3 acompanhamentos) - R$14,00",
    "Creme de açaí 500ml (3 acompanhamentos) - R$15,00",
    "Creme de açaí 700ml (3 acompanhamentos) - R$17,00",
    "Água mineral - R$2,00",
    "Água com gás - R$3,50",
    "Água saborizada - R$4,20",
    "Refrigerante mini - R$2,50",
    "Refrigerante lata - R$5,00",
    "Refrigerante Ks - R$5,00",
    "Suco Tial - R$5,00",
    "Power Ade - R$6,00",
    "Vitamil/Chocomil - R$2,50",
    "Whey - R$8,00",
    "Energético - R$13,00",
    "Coca-Cola 2L - R$13,00",
    "Sanduíche natural - R$10,00",
    "Sanduíche Siabatta - R$12,00",
    "Misto quente - R$5,00",
    "Pão de queijo recheado - R$4,50",
]

acrescimos = {
    "Leite em pó": 2.00,
    "Leite condensado": 2.00,
    "Nutella": 3.00,
    "Paçoca": 2.00
}

# Preço base do produto (a ser definido em algum lugar no script)
preco_base = 6.50  # Exemplo de valor padrão

# Dicionário para armazenar as variáveis dos acréscimos
acr_vars = {}

def salvar_dados():
    nome_cliente = nome_entry.get().strip()
    alimento_selecionado = combobox_alimentos.get().strip()
    
    if not nome_cliente or not alimento_selecionado:
        messagebox.showwarning("Campos Vazios", "Por favor, preencha todos os campos.")
        return
    
    # Parse do preço do alimento
    alimento, preco = alimento_selecionado.rsplit(' - ', 1)
    preco = float(preco.replace('R$', '').replace(',', '.'))
    
    # Adicionar acréscimos ao preço
    acr_selecionados = [acr for acr, var in acr_vars.items() if var.get()]
    acr_selecionados = list(set(acr_selecionados))  # Remover duplicatas
    total_acrescimos = sum(acrescimos[acr] for acr in acr_selecionados)
    preco_total = preco_base + total_acrescimos
    
    # Nome da pasta e arquivos
    pasta_clientes = 'CLIENTES'
    nome_arquivo_xlsx = os.path.join(pasta_clientes, f'{nome_cliente}.xlsx')
    nome_arquivo_txt = os.path.join(pasta_clientes, f'{nome_cliente}.txt')

    # Criar a pasta CLIENTES se não existir
    if not os.path.exists(pasta_clientes):
        os.makedirs(pasta_clientes)
    
    # Verificar se o Excel está instalado
    excel_instalado = False
    try:
        import win32com.client
        excel_instalado = True
    except ImportError:
        pass
    
    if excel_instalado:
        # Usar Excel
        if os.path.exists(nome_arquivo_xlsx):
            workbook = load_workbook(nome_arquivo_xlsx)
        else:
            workbook = Workbook()
        
        sheet = workbook.active
        sheet.title = nome_cliente
        
        if sheet.max_row == 1:
            sheet.append(["🕛 Data e Hora", "🍔 Alimento", "💲 Preço", "📦 Acréscimos"])
        
        if sheet["A" + str(sheet.max_row)].value == "Total Gasto":
            sheet.delete_rows(sheet.max_row, 1)
        
        data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        acr_texto = ', '.join(acr_selecionados) if acr_selecionados else "Nenhum"
        sheet.append([data_hora, alimento, preco_total, acr_texto])
        
        total_gasto = sum(sheet.cell(row=i, column=3).value for i in range(2, sheet.max_row + 1))
        sheet.append(["Total Gasto", "", total_gasto, ""])
        
        workbook.save(nome_arquivo_xlsx)
        messagebox.showinfo("Dados Salvos", f"Dados de {nome_cliente} salvos com sucesso na planilha!")
    
    else:
        # Usar Bloco de Notas
        if os.path.exists(nome_arquivo_txt):
            with open(nome_arquivo_txt, 'r') as file:
                linhas = file.readlines()
        else:
            linhas = ["Data e Hora\tAlimento\tPreço\tAcréscimos\n"]
        
        data_hora = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        acr_texto = ', '.join(acr_selecionados) if acr_selecionados else "Nenhum"
        nova_linha = f"{data_hora}\t{alimento}\tR${preco_total:.2f}\t{acr_texto}\n"
        linhas.append(nova_linha)
        
        # Calcular o total gasto
        total_gasto = calcular_total_gasto(linhas)
        
        # Atualizar ou adicionar linha do total gasto
        if linhas[-1].startswith("Total Gasto"):
            linhas[-1] = f"Total Gasto\t\tR${total_gasto:.2f}\t\n"
        else:
            linhas.append(f"Total Gasto\t\tR${total_gasto:.2f}\t\n")
        
        with open(nome_arquivo_txt, 'w') as file:
            file.writelines(linhas)
        
        messagebox.showinfo("Dados Salvos", f"Dados de {nome_cliente} salvos com sucesso no arquivo de texto!")

    # Limpar campos
    nome_entry.delete(0, ctk.END)
    combobox_alimentos.set('')
    for var in acr_vars.values():
        var.set(0)

def calcular_total_gasto(linhas):
    total = 0.0
    for linha in linhas[1:]:
        partes = linha.split('\t')
        if len(partes) > 2 and partes[2]:
            total += float(partes[2].replace('R$', '').replace(',', '.'))
    return total

def atualizar_preco():
    acr_selecionados = [acr for acr, var in acr_vars.items() if var.get()]
    acr_selecionados = list(set(acr_selecionados))  # Remover duplicatas
    total_acrescimos = sum(acrescimos[acr] for acr in acr_selecionados)
    preco_total = preco_base + total_acrescimos
    preco_total_label.configure(text=f"Preço Total: R${preco_total:.2f}")

# Configuração inicial do customtkinter
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

# Verificar a data atual
data_limite = datetime(2024, 8, 5)
data_atual = datetime.now()

if data_atual > data_limite:
    messagebox.showerror("Período de Teste Esgotado")
else:
    # Criação da janela principal
    janela = ctk.CTk()
    janela.title("Registro de Consumo")
    janela.geometry("600x600")

    # Adicionar estilo
    janela.configure(bg='#1a1a1a')

    # Desabilitar redimensionamento da janela
    janela.resizable(False, False)

    # Frame para a logo e título
    frame_logo = ctk.CTkFrame(janela, fg_color="transparent")
    frame_logo.pack(pady=20)


    # Título
    titulo_label = ctk.CTkLabel(frame_logo, text="Registro de Consumo", font=ctk.CTkFont(size=24, weight="bold"))
    titulo_label.pack(pady=10)

    # Frame para os campos de entrada
    frame_entrada = ctk.CTkFrame(janela)
    frame_entrada.pack(pady=10)

    # Campo para nome do cliente
    nome_label = ctk.CTkLabel(frame_entrada, text="Nome do Cliente:")
    nome_label.grid(row=0, column=0, padx=10, pady=5, sticky="w")
    nome_entry = ctk.CTkEntry(frame_entrada, width=300)
    nome_entry.grid(row=0, column=1, padx=10, pady=5)

    # Combobox para selecionar o alimento
    alimento_label = ctk.CTkLabel(frame_entrada, text="Selecione o Alimento:")
    alimento_label.grid(row=1, column=0, padx=10, pady=5, sticky="w")
    combobox_alimentos = ctk.CTkComboBox(frame_entrada, values=alimentos, width=300)
    combobox_alimentos.grid(row=1, column=1, padx=10, pady=5)

    # Frame para os acréscimos
    frame_acrescimos = ctk.CTkFrame(janela)
    frame_acrescimos.pack(pady=10)

    acrescimos_label = ctk.CTkLabel(frame_acrescimos, text="Acréscimos:")
    acrescimos_label.grid(row=0, column=0, padx=10, pady=5, sticky="w")

    for idx, (acr, preco) in enumerate(acrescimos.items()):
        acr_vars[acr] = ctk.BooleanVar()
        acr_checkbox = ctk.CTkCheckBox(frame_acrescimos, text=f"{acr} - R${preco:.2f}", variable=acr_vars[acr], command=atualizar_preco)
        acr_checkbox.grid(row=idx+1, column=0, padx=10, pady=5, sticky="w")

    # Preço total
    preco_total_label = ctk.CTkLabel(janela, text="Preço Total: R$0.00", font=ctk.CTkFont(size=16, weight="bold"))
    preco_total_label.pack(pady=10)

    # Botão para salvar
    salvar_button = ctk.CTkButton(janela, text="Salvar", command=salvar_dados)
    salvar_button.pack(pady=20)

    # Rodapé
    rodape_label = ctk.CTkLabel(janela, text="© 2024 Sua Empresa - Todos os direitos reservados", font=ctk.CTkFont(size=10))
    rodape_label.pack(side="bottom", pady=10)

    janela.mainloop()